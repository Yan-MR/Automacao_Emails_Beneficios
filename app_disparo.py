import os
import sys
import json
import threading
import time
import shutil # --- IMPORTAÇÃO PARA O BACKUP ---
import pythoncom
import win32com.client as win32
import re
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from PIL import Image, ImageTk

# --- FUNÇÃO MÁGICA PARA A LOGO NO .EXE ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- CONFIGURAÇÕES DE DIRETÓRIO DINÂMICAS ---
if getattr(sys, 'frozen', False):
    PASTA_BASE = os.path.dirname(sys.executable)
else:
    PASTA_BASE = os.path.dirname(os.path.abspath(__file__))

ARQUIVO_CONFIG = os.path.join(PASTA_BASE, "config_email_beneficios.json")

# 6 Arquivos de Template exclusivos para o E-mail
ARQUIVOS_TEMPLATES = {
    "sedex": os.path.join(PASTA_BASE, "email_template_sedex.txt"),
    "scs": os.path.join(PASTA_BASE, "email_template_scs.txt"),
    "berrini": os.path.join(PASTA_BASE, "email_template_berrini.txt"),
    "reaviso_scs": os.path.join(PASTA_BASE, "email_template_reaviso_scs.txt"),
    "reaviso_berrini": os.path.join(PASTA_BASE, "email_template_reaviso_berrini.txt"),
    "personalizada": os.path.join(PASTA_BASE, "email_template_personalizada.txt") # --- NOVO TEMPLATE ---
}

TEXTOS_PADROES = {
    "sedex": """Boa tarde.
Olá {primeiro_nome}, tudo bem?

É um prazer tê-lo em nossa CIA como um de nossos (as) colaboradores (as), nós do time de benefícios temos uma excelente noticia para você!
Informamos que o seu cartão *Alelo*, foi entregue em nosso HUB SCS e iremos redireciona-lo via SEDEX ao endereço cadastrado em sistema.

Segue abaixo código de rastreio, para acessar, basta entrar no link: https://www.correios.com.br/

[TABELA_RASTREIO]

Obs.: Esta é uma mensagem automática. Por favor não responda este e-mail.

Atenciosamente,
*Grupo Casas Bahia - Gente, Gestão e Sustentabilidade*
Operações de Benefícios""",

    "scs": """Olá {primeiro_nome}, tudo bem?
    
Informamos que o seu *Cartão Alelo* foi recebido em nosso Hub e está disponível para a retirada.

📌 *Atenção colaboradores alocados sistemicamente no Hub SCS – Filial 1580, segue abaixo informações de retirada.*

📍 *Local de retirada:*
*Hub SCS - 2° andar – Mesa n° 636 (Terceira baia a frente das salas de reuniões.)*

🕒 *Horários de atendimento:*
SEG | QUA | QUI | SEX (Sem entregas as terças-feiras)
*Manhã:* das 9h às 11h30
*Tarde:* das 12h30 às 16h

Pedimos atenção à data para garantir o recebimento do cartão no local indicado.

❓ *Dúvidas frequentes:*

*• Onde identificar minha Filial de cadastro?*
Para consulta sua filial de cadastro, acesse o Portal do Colaborador> Meu perfil> a informação estará abaixo do seu nome, departamento e diretoria.

*• Posso solicitar que outra pessoa retire meu cartão?*
Sim. A retirada pode ser feita por outro colaborador, desde que informe seu *nome completo e matrícula* no momento da retirada.

Atenciosamente;
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "berrini": """Olá {primeiro_nome}, tudo bem?
    
Informamos que o seu *Cartão Alelo* foi recebido em nosso Hub e encontra-se disponível para retirada conforme as orientações abaixo.

📌 *Colaboradores alocados sistemicamente na Estação Casas Bahia (Berrini) – Filiais 01 | 650 | 1968:*
A retirada deve ocorrer no dia *27/02/2026*, de forma pontual, conforme a ida de um de nossos portadores ao local.

📍 *Local de retirada:*
*Hub Estação Casas Bahia (Berrini) - 4º andar - sala de Bem-Estar (ao lado do Espaço Viver Bem)*

🕒 *Horário de atendimento:*
9h30 às 17h

Pedimos atenção a data informada referente ao plantão de entrega.

❓ *Dúvidas frequentes:*

*• Onde identificar minha Filial de cadastro?*
Para consulta sua filial de cadastro, acesse o Portal do Colaborador> Meu perfil> a informação estará abaixo do seu nome, departamento e diretoria.

*• Posso solicitar que outra pessoa retire meu cartão?*
Sim. A retirada pode ser feita por outro colaborador, desde que informe seu *nome completo e matrícula* no momento da retirada.
Caso já tenha retirado o seu cartão Alelo, por favor desconsiderar este e-mail!!

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "reaviso_scs": """Olá {primeiro_nome}, tudo bem?
 
Identificamos que você ainda não realizou a retirada do seu Cartão Alelo.

Gostaríamos de lembrá-lo(a) que o seu cartão está disponível para retirada hoje, 04/03/2026.
 
📍 *Local de retirada:*
Hub SCS – 2º andar – Mesa 636.
 
🕒 *Horário de atendimento:*
09h às 15h.

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    "reaviso_berrini": """Olá {primeiro_nome}, tudo bem?
 
Identificamos que você ainda não realizou a retirada do seu Cartão Alelo.

Gostaríamos de lembrá-lo(a) que o seu cartão está disponível para retirada hoje, 04/03/2026.
 
📍 *Local de retirada:*
Hub Estação Casas Bahia (Berrini) - 4º andar - sala de Bem-Estar (ao lado do Espaço Viver Bem)
 
🕒 *Horário de atendimento:*
09h às 15h.

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br""",

    # --- NOVO TEXTO PADRÃO PARA MENSAGEM PERSONALIZADA ---
    "personalizada": """Olá {primeiro_nome}, tudo bem?

[Apague este texto e digite o seu comunicado aqui. Você pode usar a tag {primeiro_nome} quantas vezes quiser no texto para o robô trocar pelo nome da pessoa.]

Atenciosamente; 
*Operações Benefícios*
adm.beneficios@casasbahia.com.br"""
}

# --- FUNÇÕES DE SUPORTE ---
def carregar_template(tipo):
    caminho = ARQUIVOS_TEMPLATES[tipo]
    if os.path.exists(caminho):
        try:
            with open(caminho, "r", encoding="utf-8") as f: return f.read()
        except Exception: return TEXTOS_PADROES[tipo]
    else:
        with open(caminho, "w", encoding="utf-8") as f: f.write(TEXTOS_PADROES[tipo])
        return TEXTOS_PADROES[tipo]

def salvar_template(tipo, texto):
    with open(ARQUIVOS_TEMPLATES[tipo], "w", encoding="utf-8") as f: f.write(texto)


# --- LÓGICA DO ROBÔ DE E-MAIL ---
def processar_disparos_email(app_gui, caminho_arquivo, template_texto_puro):
    # Necessário para o Outlook rodar em uma Thread secundária no Windows
    pythoncom.CoInitialize() 
    
    app_gui.atualizar_status("Lendo planilha e conectando ao Outlook...", INFO)
    app_gui.btn_iniciar.config(state="disabled")
    
    try:
        wb = load_workbook(caminho_arquivo)
        sh_ROBO = wb["ROBO"]
        
        # Pega as configurações de C4 e C6, garantindo que não deem erro se estiverem vazias
        assunto = sh_ROBO["C4"].value if sh_ROBO["C4"].value else "Atualização - Cartão Alelo"
        bcc = sh_ROBO["C6"].value
        remetente_oficial = "adm.beneficios@casasbahia.com.br" 
        
        try:
            outlook = win32.Dispatch('outlook.application')
        except Exception as e:
            app_gui.atualizar_status("Erro: O Outlook não está aberto ou configurado.", DANGER)
            app_gui.btn_iniciar.config(state="normal")
            return

        df_completo = pd.read_excel(caminho_arquivo, sheet_name="ROBO", header=7)
        df_completo.columns = df_completo.columns.str.strip()
        
        df_pendentes = df_completo[(df_completo['Status'] != 'Enviado') & (df_completo['Enviar'] == 'x')]
        
        if df_pendentes.empty:
            app_gui.atualizar_status("Nenhuma mensagem pendente na planilha!", WARNING)
            app_gui.btn_iniciar.config(state="normal")
            return

        total = len(df_pendentes)
        enviados = 0
        lista_falhas = []
        
        for index, row in df_pendentes.iterrows():
            email_destino = str(row.get('Email', '')).strip()
            
            # --- BLINDAGEM CONTRA LINHAS VAZIAS ---
            if not email_destino or email_destino.lower() == 'nan':
                continue
                
            nome = str(row.get('Nome', 'Colaborador')).strip()
            rastreio = str(row.get('Código de Rastreio', 'N/D'))
            matricula = str(row.get('Matricula', 'N/D')).split('.')[0] 
            cargo = str(row.get('Cargo', 'N/D'))
            
            # Tratamento da data para não vir como Timestamp do Pandas
            try:
                data_postagem = pd.to_datetime(row.get('Data de Postagem')).strftime('%d/%m/%Y')
            except:
                data_postagem = str(row.get('Data de Postagem', 'N/D')).split(' ')[0]
            
            primeiro_nome = nome.split()[0].capitalize() if nome and nome.lower() != 'nan' else "Colaborador"
            
            # Conversão do texto puro para HTML do Outlook
            texto_preparado = template_texto_puro.replace("{primeiro_nome}", primeiro_nome)
            html_miolo = texto_preparado.replace('\n', '<br>')
            html_miolo = re.sub(r'\*(.*?)\*', r'<b>\1</b>', html_miolo) # Negrito
            
            if "[TABELA_RASTREIO]" in html_miolo:
                tabela_html = f"""
                <br>
                <table style="border-collapse: collapse; width: 100%; max-width: 900px; font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000; text-align: center; margin-top: 15px; margin-bottom: 15px;">
                    <thead>
                        <tr>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-weight: bold;">Matricula</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-weight: bold;">Nome</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-weight: bold;">Cargo</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-weight: bold;">Código de Rastreio</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-weight: bold;">Data de Postagem</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td style="border: 1px solid black; padding: 8px;">{matricula}</td>
                            <td style="border: 1px solid black; padding: 8px;">{nome}</td>
                            <td style="border: 1px solid black; padding: 8px;">{cargo}</td>
                            <td style="border: 1px solid black; padding: 8px;">{rastreio}</td>
                            <td style="border: 1px solid black; padding: 8px;">{data_postagem}</td>
                        </tr>
                    </tbody>
                </table>
                <br>"""
                html_miolo = html_miolo.replace("[TABELA_RASTREIO]", tabela_html)
            
            # Envelopa tudo com a fonte oficial do E-mail
            html_final_email = f"""
            <html>
                <body style="font-family: Calibri, Arial, sans-serif; color: #1F3864; font-size: 11pt; line-height: 1.5;">
                    {html_miolo}
                </body>
            </html>
            """
            
            app_gui.atualizar_status(f"🚀 Enviando E-mail para: {primeiro_nome}... ({enviados+1}/{total})", PRIMARY)
            
            linha_excel = index + 9
            coluna_status = 7
            for col in range(1, sh_ROBO.max_column + 1):
                if str(sh_ROBO.cell(row=8, column=col).value).strip() == "Status": 
                    coluna_status = col
                    break

            try:
                # Disparo pelo Outlook COM
                mail = outlook.CreateItem(0)
                mail.SentOnBehalfOfName = remetente_oficial
                mail.Subject = str(assunto)
                mail.To = email_destino
                if bcc:
                    mail.BCC = str(bcc)
                mail.HTMLBody = html_final_email
                
                # --- O TRUQUE DE MESTRE PARA SALVAR NA CAIXA CERTA ---
                try:
                    remetente_obj = outlook.Session.CreateRecipient(remetente_oficial)
                    remetente_obj.Resolve() 
                    
                    if remetente_obj.Resolved:
                        pasta_enviados_compartilhada = outlook.Session.GetSharedDefaultFolder(remetente_obj, 5)
                        mail.SaveSentMessageFolder = pasta_enviados_compartilhada
                except Exception as err:
                    print(f"Aviso: Não conseguiu acessar a pasta enviados do grupo. {err}")
                # -----------------------------------------------------

                mail.Send()
                
                sh_ROBO.cell(row=linha_excel, column=coluna_status).value = "Enviado"
                wb.save(caminho_arquivo) 
                enviados += 1
                time.sleep(0.5) # Pausa rápida para não sobrecarregar o Outlook
                
            except Exception as e:
                print(f"Erro em {nome}: {e}")
                sh_ROBO.cell(row=linha_excel, column=coluna_status).value = "Não Encontrado"
                wb.save(caminho_arquivo)
                lista_falhas.append(str(nome))

        app_gui.btn_iniciar.config(state="normal")
        
        if lista_falhas:
            app_gui.atualizar_status(f"⚠️ Concluído com {len(lista_falhas)} falha(s).", WARNING)
            nomes_falha = "\n".join(lista_falhas)
            messagebox.showwarning("Atenção - Relatório de E-mails", f"Processo finalizado!\n\n✅ Sucesso: {enviados}\n❌ Falhas: {len(lista_falhas)}\n\nAs seguintes pessoas falharam (e-mail inválido/erro Outlook):\n\n{nomes_falha}")
        else:
            app_gui.atualizar_status(f"🎉 Todos os e-mails enviados com sucesso!", SUCCESS)
            messagebox.showinfo("Sucesso Total", "Todos os disparos foram concluídos perfeitamente via Outlook!")

    except Exception as e:
        app_gui.atualizar_status(f"Erro Crítico no Robô: {str(e)[:50]}", DANGER)
        app_gui.btn_iniciar.config(state="normal")
    finally:
        pythoncom.CoUninitialize()


# --- FRONT-END (INTERFACE GRÁFICA) ---
class AppEmail(tb.Window):
    def __init__(self):
        super().__init__(themename="litera") 
        self.title("Robô Envio Outlook Alelo") 
        self.geometry("1250x850") 
        self.resizable(False, False)
        
        caminho_icone = resource_path("logo.png")
        if os.path.exists(caminho_icone):
            try:
                img_icon = Image.open(caminho_icone)
                icone = ImageTk.PhotoImage(img_icon)
                self.iconphoto(False, icone)
            except Exception:
                pass 
        
        self.caminho_planilha = ""
        self.tipo_ativo = "sedex" 
        
        self.var_status = tb.StringVar(value="Pronto para iniciar.")
        self.var_tipo_msg = tb.StringVar(value="sedex")
        
        self.construir_interface()
        self.carregar_config_anterior()

    def construir_interface(self):
        frame_main = tb.Frame(self, padding=20)
        frame_main.pack(fill=BOTH, expand=True)
        
        # --- CABEÇALHO E BOTÃO DE AJUDA ---
        frame_header = tb.Frame(frame_main)
        frame_header.pack(fill=X, pady=(0, 15))
        
        frame_titles = tb.Frame(frame_header)
        frame_titles.pack(side=LEFT)
        
        lbl_titulo = tb.Label(frame_titles, text="Automação Microsoft Outlook", font=("Segoe UI", 20, "bold"), bootstyle=PRIMARY)
        lbl_titulo.pack(anchor=W)
        lbl_subtitulo = tb.Label(frame_titles, text="Módulo de Disparos por E-mail - Cartão Alelo", font=("Segoe UI", 10), foreground="gray")
        lbl_subtitulo.pack(anchor=W)
        
        btn_ajuda = tb.Button(frame_header, text="Como utilizar❓", bootstyle="info-outline", command=self.mostrar_ajuda)
        btn_ajuda.pack(side=RIGHT, anchor=N)
        # ----------------------------------
        
        frame_rodape = tb.Frame(frame_main)
        frame_rodape.pack(fill=X, side=BOTTOM)
        self.lbl_status = tb.Label(frame_rodape, textvariable=self.var_status, font=("Segoe UI", 10, "bold"), bootstyle=SECONDARY)
        self.lbl_status.pack(side=LEFT)

        self.btn_iniciar = tb.Button(frame_main, text="▶ INICIAR DISPAROS DE E-MAIL", bootstyle="success", padding=10, command=self.iniciar_disparos)
        self.btn_iniciar.pack(fill=X, side=BOTTOM, pady=(15, 15))

        frame_content = tb.Frame(frame_main)
        frame_content.pack(fill=BOTH, expand=True)

        # ====== COLUNA ESQUERDA ======
        frame_left = tb.Frame(frame_content)
        frame_left.pack(side=LEFT, fill=Y, expand=False, padx=(0, 15))

        frame_arquivo = tb.LabelFrame(frame_left, text=" 1. Base de Disparos ")
        frame_arquivo.pack(fill=X, pady=(0, 15), ipadx=5, ipady=5)
        
        frame_botoes_arquivo = tb.Frame(frame_arquivo)
        frame_botoes_arquivo.pack(anchor=W, fill=X, padx=15, pady=(10, 5))
        
        self.btn_procurar = tb.Button(frame_botoes_arquivo, text="📂 Escolher Planilha", bootstyle=SECONDARY, command=self.selecionar_arquivo)
        self.btn_procurar.pack(side=LEFT, padx=(0, 10))
        
        self.btn_limpar = tb.Button(frame_botoes_arquivo, text="🧹 Limpar Status", bootstyle="warning-outline", command=self.limpar_status_planilha)
        self.btn_limpar.pack(side=LEFT)
        
        self.lbl_caminho = tb.Label(frame_arquivo, text="Nenhuma planilha (.xlsx) selecionada", font=("Segoe UI", 9, "italic"), foreground="gray", wraplength=280)
        self.lbl_caminho.pack(anchor=W, fill=X, expand=True, padx=15, pady=(0, 10))
        
        lbl_info_assunto = tb.Label(frame_arquivo, text="O Assunto e BCC são lidos\ndas células C4 e C6 da aba ROBO.", font=("Segoe UI", 8), foreground="gray")
        lbl_info_assunto.pack(anchor=W, padx=15, pady=(0, 10))

        frame_tipo = tb.LabelFrame(frame_left, text=" 2. Tipo de Comunicação ")
        frame_tipo.pack(fill=X, pady=(0, 0), ipadx=5, ipady=5)
        tb.Radiobutton(frame_tipo, text="Envio Correios (Sedex)", variable=self.var_tipo_msg, value="sedex", command=self.trocar_aba).pack(anchor=W, padx=15, pady=(10, 5))
        tb.Radiobutton(frame_tipo, text="Retirada Presencial (Hub SCS)", variable=self.var_tipo_msg, value="scs", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Retirada Presencial (Hub Berrini)", variable=self.var_tipo_msg, value="berrini", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Separator(frame_tipo).pack(fill=X, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Re-aviso (Hub SCS)", variable=self.var_tipo_msg, value="reaviso_scs", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Radiobutton(frame_tipo, text="Re-aviso (Hub Berrini)", variable=self.var_tipo_msg, value="reaviso_berrini", command=self.trocar_aba).pack(anchor=W, padx=15, pady=5)
        tb.Separator(frame_tipo).pack(fill=X, padx=15, pady=5)
        # --- O NOVO BOTÃO DE MENSAGEM PERSONALIZADA ---
        tb.Radiobutton(frame_tipo, text="Mensagem Personalizada (Livre)", variable=self.var_tipo_msg, value="personalizada", command=self.trocar_aba).pack(anchor=W, padx=15, pady=(5, 10))

        # ====== COLUNA DIREITA ======
        frame_right = tb.Frame(frame_content)
        frame_right.pack(side=LEFT, fill=BOTH, expand=True)

        frame_msg = tb.LabelFrame(frame_right, text=" 3. Pré-visualização e Edição do E-mail (Texto Simples) ")
        frame_msg.pack(fill=BOTH, expand=True, ipadx=5, ipady=5)
        
        lbl_dica = tb.Label(frame_msg, text="💡 DICA: Escreva normalmente. Para negrito, use *asteriscos*. O robô ajusta as a formatação automaticamente.", font=("Segoe UI", 9, "bold"), bootstyle=INFO)
        lbl_dica.pack(anchor=W, padx=15, pady=(5,0))
        
        scroll_txt = tb.Scrollbar(frame_msg)
        scroll_txt.pack(side=RIGHT, fill=Y, pady=5, padx=(0,10))
        self.txt_mensagem = tb.Text(frame_msg, font=("Segoe UI", 10), wrap="word", yscrollcommand=scroll_txt.set)
        self.txt_mensagem.pack(side=LEFT, fill=BOTH, expand=True, padx=(15,0), pady=5)
        scroll_txt.config(command=self.txt_mensagem.yview)
        
        self.txt_mensagem.insert("1.0", carregar_template("sedex"))

    def mostrar_ajuda(self):
        texto_ajuda = (
            "PASSO A PASSO DE USO:\n\n"
            "1. PLANILHA DE DADOS:\n"
            "Clique em 'Escolher Planilha' e selecione a sua base. A planilha precisa ter a coluna 'Enviar' marcada com 'x'. (O robô cria uma cópia de segurança automática para não corromper sua base original!).\n\n"
            "2. ASSUNTO E CÓPIA (BCC):\n"
            "O robô lê o Assunto e os e-mails de Cópia Oculta diretamente das células C4 e C6 da aba ROBO no Excel.\n\n"
            "3. MENSAGEM:\n"
            "Escolha o tipo de disparo. Você pode editar o texto na tela preta. Se usar a opção 'Mensagem Personalizada', pode escrever o que quiser e usar a tag {primeiro_nome} para o robô chamar o colaborador pelo nome.\n\n"
            "4. EXECUTAR:\n"
            "Clique em 'Iniciar Disparos de E-mail'. O Outlook enviará as mensagens em segundo plano. Dica: Use 'Limpar Status' para apagar os registros da planilha e testar novamente."
        )
        messagebox.showinfo("Guia Rápido - Robô do Outlook", texto_ajuda)

    def trocar_aba(self):
        texto_atual = self.txt_mensagem.get("1.0", tk.END).strip()
        salvar_template(self.tipo_ativo, texto_atual)
        novo_tipo = self.var_tipo_msg.get()
        self.tipo_ativo = novo_tipo
        self.txt_mensagem.delete("1.0", tk.END)
        self.txt_mensagem.insert("1.0", carregar_template(novo_tipo))

    def atualizar_status(self, texto, estilo=INFO):
        self.var_status.set(texto)
        self.lbl_status.config(bootstyle=estilo)
        self.update_idletasks()

    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(title="Selecione a Planilha", filetypes=[("Planilhas do Excel", "*.xlsx")])
        if arquivo:
            self.caminho_planilha = arquivo
            self.lbl_caminho.config(text=os.path.basename(arquivo), foreground="black", font=("Segoe UI", 10, "bold"))
            self.atualizar_status("Planilha vinculada com sucesso.", INFO)
            self.salvar_config_anterior()

    def limpar_status_planilha(self):
        if not self.caminho_planilha:
            messagebox.showwarning("Aviso", "Por favor, selecione a planilha de pendências primeiro.")
            return

        resposta = messagebox.askyesno("Limpar Status", "Tem certeza que deseja apagar todos os registros 'Enviado' e 'Não Encontrado' da planilha?\n\nIsso permitirá que o robô envie e-mails novamente.")
        if not resposta: return

        try:
            wb = load_workbook(self.caminho_planilha)
            sh_ROBO = wb["ROBO"]
            coluna_status = 7
            for col in range(1, sh_ROBO.max_column + 1):
                if str(sh_ROBO.cell(row=8, column=col).value).strip() == "Status":
                    coluna_status = col
                    break
            
            linhas_apagadas = 0
            for row in range(9, sh_ROBO.max_row + 1):
                valor = str(sh_ROBO.cell(row=row, column=coluna_status).value)
                if valor in ["Enviado", "Não Encontrado"]:
                    sh_ROBO.cell(row=row, column=coluna_status).value = ""
                    linhas_apagadas += 1
                    
            wb.save(self.caminho_planilha)
            messagebox.showinfo("Sucesso", f"Status limpos com sucesso!\n{linhas_apagadas} linhas foram resetadas.")
            self.atualizar_status(f"Planilha resetada ({linhas_apagadas} linhas). Pronto para envio de E-mails.", SUCCESS)
            
        except PermissionError:
            messagebox.showerror("Erro de Permissão", "A planilha está aberta no Excel! Feche o arquivo e tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao limpar a planilha:\n{e}")

    def carregar_config_anterior(self):
        if os.path.exists(ARQUIVO_CONFIG):
            try:
                with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
                    dados = json.load(f)
                    caminho_salvo = dados.get("caminho", "")
                    if os.path.exists(caminho_salvo):
                        self.caminho_planilha = caminho_salvo
                        self.lbl_caminho.config(text=os.path.basename(caminho_salvo), foreground="black", font=("Segoe UI", 10, "bold"))
            except Exception: pass

    def salvar_config_anterior(self):
        try:
            with open(ARQUIVO_CONFIG, 'w', encoding='utf-8') as f:
                json.dump({"caminho": self.caminho_planilha}, f)
        except Exception: pass

    def iniciar_disparos(self):
        if not self.caminho_planilha:
            messagebox.showwarning("Aviso", "Por favor, selecione a planilha de pendências antes de iniciar o envio.")
            return
            
        # --- A MÁGICA DA CÓPIA DE SEGURANÇA NO E-MAIL ---
        diretorio = os.path.dirname(self.caminho_planilha)
        nome_arquivo = os.path.basename(self.caminho_planilha)
        
        if not nome_arquivo.endswith("_Copia_Segura.xlsx"):
            nome_sem_ext, ext = os.path.splitext(nome_arquivo)
            caminho_copia = os.path.join(diretorio, f"{nome_sem_ext}_Copia_Segura{ext}")
            
            try:
                shutil.copy2(self.caminho_planilha, caminho_copia)
                self.caminho_planilha = caminho_copia
                self.lbl_caminho.config(text=os.path.basename(caminho_copia), foreground="blue")
                self.atualizar_status(f"Cópia de segurança criada e em uso.", INFO)
            except Exception as e:
                messagebox.showerror("Erro de Cópia", f"Não foi possível criar a cópia de segurança:\n{e}")
                return
        # ------------------------------------------------
            
        texto_atual = self.txt_mensagem.get("1.0", tk.END).strip()
        salvar_template(self.tipo_ativo, texto_atual) 
        
        if self.tipo_ativo == "sedex" and "[TABELA_RASTREIO]" not in texto_atual:
            if not messagebox.askyesno("Cuidado!", "Você apagou a tag [TABELA_RASTREIO] do Sedex. A tabela não será gerada.\nEnviar mesmo assim?"): return
        elif "{primeiro_nome}" not in texto_atual:
            if not messagebox.askyesno("Cuidado!", "Você apagou a tag {primeiro_nome}. A mensagem não será personalizada.\nEnviar mesmo assim?"): return

        threading.Thread(target=processar_disparos_email, args=(self, self.caminho_planilha, texto_atual), daemon=True).start()

if __name__ == "__main__":
    app = AppEmail()
    app.mainloop()
    