import win32com.client as win32
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime

def enviar_emails_beneficios():
    # 1. CONFIGURAÇÕES INICIAIS
    # Mude apenas o nome do arquivo abaixo para o nome real da sua planilha
    caminho_arquivo = "Template_Beneficios.xlsx"
    
    # Carrega o arquivo com openpyxl (usado no final para escrever "Enviado")
    wb = load_workbook(caminho_arquivo)
    sh_capa = wb["Capa"]
    
    # Pega Assunto e Cópia Oculta direto da planilha
    assunto = sh_capa["C4"].value
    bcc = sh_capa["C6"].value
    remetente_oficial = "rodriguesyan143@gmail.com"
    
    # Inicia a conexão com o Outlook (Precisa estar aberto no PC)
    outlook = win32.Dispatch('outlook.application')
    emails_enviados = 0
    
    # 2. LENDO OS DADOS (Aba 'Capa', cabeçalho na linha 8 do Excel)
    df_completo = pd.read_excel(caminho_arquivo, sheet_name="Capa", header=7)
    
    # Filtra apenas quem não tem "Enviado" e está com o "x" marcado
    # Se na sua planilha estiver escrito "E-mail" em vez de "Email", ajuste aqui e lá embaixo
    df_pendentes = df_completo[
        (df_completo['Status'] != 'Enviado') & 
        (df_completo['Enviar'] == 'x')
    ]
    
    # Agrupa pelas pessoas para não mandar e-mail duplicado
    grupos = df_pendentes.groupby('Nome') 
    
    # 3. LOOP DE DISPARO
    for nome_agrupado, dados_responsavel in grupos:
        
        # Pega o e-mail de destino da coluna respectiva
        destino = dados_responsavel['Email'].iloc[0] 
        
        # --- DECLARANDO AS VARIÁVEIS PARA A TABELA ---
        # Esses nomes entre [' '] precisam ser EXATAMENTE os do cabeçalho do Excel
        matricula = dados_responsavel['Matricula'].iloc[0]
        nome_colaborador = dados_responsavel['Nome'].iloc[0]
        cargo = dados_responsavel['Cargo'].iloc[0]
        rastreio = dados_responsavel['Código de Rastreio'].iloc[0]
        
        # Formatando a data de postagem para o padrão DD/MM/AAAA
        data_bruta = pd.to_datetime(dados_responsavel['Data de Postagem'].iloc[0])
        data_postagem = data_bruta.strftime('%d/%m/%Y')
        
        # --- MONTANDO O CORPO DO E-MAIL (HTML + CSS) ---
        corpo_html = f"""
        <html>
            <body style="font-family: Calibri, Arial, sans-serif; color: #1F3864; font-size: 11pt; line-height: 1.5;">
                <p>Boa tarde.</p>
                
                <p>Tudo bem?</p>
                
                <p>É um prazer tê-lo em nossa CIA como um de nossos (as) colaboradores (as), nós do time de benefícios temos uma excelente noticia para você!<br>
                Informamos que o seu cartão <strong>Alelo</strong>, foi entregue em nosso HUB SCS e iremos redireciona-lo via SEDEX ao endereço cadastrado em sistema.<br>
                Segue abaixo código de rastreio, para acessar, basta entrar no link: <a href="https://www.correios.com.br/" style="color: #0563C1; text-decoration: underline;">https://www.correios.com.br/</a></p>

                <table style="border-collapse: collapse; width: 100%; max-width: 900px; font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000; text-align: center; margin-top: 20px; margin-bottom: 20px;">
                    <thead>
                        <tr>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-style: italic; font-weight: bold;">Matricula</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-style: italic; font-weight: bold;">Nome</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-style: italic; font-weight: bold;">Cargo</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-style: italic; font-weight: bold;">Código de Rastreio</th>
                            <th style="border: 1px solid black; background-color: #9BC2E6; padding: 8px; font-style: italic; font-weight: bold;">Data de Postagem</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td style="border: 1px solid black; padding: 8px;">{matricula}</td>
                            <td style="border: 1px solid black; padding: 8px;">{nome_colaborador}</td>
                            <td style="border: 1px solid black; padding: 8px;">{cargo}</td>
                            <td style="border: 1px solid black; padding: 8px;">{rastreio}</td>
                            <td style="border: 1px solid black; padding: 8px;">{data_postagem}</td>
                        </tr>
                    </tbody>
                </table>

                <p style="color: #1F3864;">Obs.: Esta é uma mensagem automática. Por favor não responda este e-mail</p>

                <p style="color: #1F3864; margin-top: 30px;">Atenciosamente,<br>
                Grupo Casas Bahia - Gente, Gestão e Sustentabilidade<br>
                Operações de Benefícios</p>
            </body>
        </html>
        """
        
        # --- PREPARANDO O OUTLOOK ---
        mail = outlook.CreateItem(0)
        mail.SentOnBehalfOfName = remetente_oficial
        mail.Subject = assunto
        mail.To = destino
        if bcc:
            mail.BCC = bcc
        
        mail.HTMLBody = corpo_html
        
        # ENVIAR
        mail.Send()
        emails_enviados += 1
        
        # --- ATUALIZANDO O STATUS NA PLANILHA ---
        indices_originais = dados_responsavel.index
        for idx in indices_originais:
            linha_excel = idx + 9 # Pandas começa no 0, dados do Excel na linha 9
            sh_capa.cell(row=linha_excel, column=7).value = "Enviado" # Coluna G (7)
            
    # 4. SALVAR E FINALIZAR
    wb.save(caminho_arquivo)
    print(f"Processo concluído! E-mails enviados: {emails_enviados:03d}")

if __name__ == "__main__":
    enviar_emails_beneficios()