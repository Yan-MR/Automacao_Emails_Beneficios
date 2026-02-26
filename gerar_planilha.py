from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def criar_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Capa"

    # 1. Configurações de E-mail
    ws["B4"] = "Assunto:"
    ws["C4"] = "Entrega de Cartão Alelo - Atualização de Rastreio"
    ws["B6"] = "Cópia Oculta (BCC):"
    ws["C6"] = "" # Deixe em branco se não quiser mandar cópia agora

    # 2. Cabeçalhos na Linha 8
    headers = {
        "A": "ID", 
        "B": "Matricula", 
        "C": "Nome", 
        "D": "Cargo", 
        "E": "Código de Rastreio", 
        "F": "Data de Postagem", 
        "G": "Status", 
        "H": "Obs", 
        "I": "Email", 
        "J": "Enviar"
    }

    # Estilo do cabeçalho (Azul escuro com letra branca)
    fundo_azul = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    letra_branca = Font(bold=True, color="FFFFFF")

    for col, nome in headers.items():
        celula = ws[f"{col}8"]
        celula.value = nome
        celula.font = letra_branca
        celula.fill = fundo_azul

    # 3. Linha 9 - Dados de Teste (Baseados no seu print)
    dados_exemplo = {
        "A": 1,
        "B": 5092930,
        "C": "MARCOS DIOGO OLIVEIRA RODRIGUES",
        "D": "SOFTWARE ENGINEERING SPECIALIST II - OFFICELESS",
        "E": "AD107082205BR",
        "F": "20/02/2026",
        "G": "", # Status vazio para a automação poder rodar
        "H": "Teste",
        "I": "rodriguesyan143@gmail.com", # Seu email para você receber o teste
        "J": "x" # O 'x' que engatilha o envio no seu código
    }

    for col, valor in dados_exemplo.items():
        ws[f"{col}9"] = valor

    # Ajusta a largura das colunas principais para ficar mais bonito
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['I'].width = 30

    wb.save("Template_Beneficios.xlsx")
    print("Pronto! Arquivo 'Template_Beneficios.xlsx' gerado na pasta.")

if __name__ == "__main__":
    criar_template()